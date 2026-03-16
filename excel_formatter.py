from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


def format_excel(file_path):

    wb = load_workbook(file_path)

    company_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
    valuation_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    quality_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    growth_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    investment_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    scoring_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for sheet in ["Watchlist", "Full Scoring"]:

        ws = wb[sheet]

        ws.insert_rows(1)

        headers = [
            "Company Data", "Company Data", "Company Data", "Company Data",
            "Valuation", "Valuation",
            "Quality", "Quality",
            "Growth",
            "Investment View", "Investment View", "Investment View",
            "Scoring", "Scoring"
        ]

        for col, value in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = value

        ws.merge_cells("A1:D1")
        ws.merge_cells("E1:F1")
        ws.merge_cells("G1:H1")
        ws.merge_cells("I1:I1")
        ws.merge_cells("J1:L1")
        ws.merge_cells("M1:N1")

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for col in range(1, 5):
            ws.cell(row=1, column=col).fill = company_fill

        for col in range(5, 7):
            ws.cell(row=1, column=col).fill = valuation_fill

        for col in range(7, 9):
            ws.cell(row=1, column=col).fill = quality_fill

        ws.cell(row=1, column=9).fill = growth_fill

        for col in range(10, 13):
            ws.cell(row=1, column=col).fill = investment_fill

        for col in range(13, 15):
            ws.cell(row=1, column=col).fill = scoring_fill

        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        score_column = None

        for col in range(1, ws.max_column + 1):
            if ws.cell(row=2, column=col).value == "total_score":
                score_column = col

        if score_column:

            col_letter = get_column_letter(score_column)

            ws.conditional_formatting.add(
                f"{col_letter}3:{col_letter}{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="E7E6E6",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="9DC3E6",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column_cells
            )

            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    ws = wb["Methodology"]

    ws.column_dimensions["A"].width = 140

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(file_path)